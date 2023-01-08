#2022-11-26

from urllib import parse
import pandas as pd
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
import threading

# 상수 선언
## 할인율
DISCOUNT_RATE_LOC: int = 87 #BBB- 5y
KIS_LINK: str = 'https://www.kisrating.com/ratingsStatistics/statics_spread.do'
## 다트 재무제표
DART_API_KEY = 'b64695f3f2a79d07bde772ffa630f935e0d050c0'
STOCKS_DB_LOC = '/workspace/PythonTrader/stocksDB.xlsx'
## Date Frame
DATE_FRAME = ['2019/12', '2020/12', '2021/12']

def getDiscountRate(link, location) -> float:
    html = requests.get(link).text
    soup = BeautifulSoup(html, "html.parser")
    result = float(soup.select('td.ar.pr40')[location].text)
    return result

# 주식 DB 읽기
workbook = openpyxl.load_workbook(STOCKS_DB_LOC)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
workbook.close()

## 주식 명으로 주식코드 가져오기
def get_code_by_name():
    target_name = input("주식명: ").upper()
    for row in range(2, rows + 1):
        comp_name = sheet.cell(row, 2).value.upper()
        if comp_name == target_name:
            return sheet.cell(row, 1).value

def check_data_date(this_obj):
    if type(this_obj) == list:
        if this_obj[0] == DATE_FRAME[0] and this_obj[1] == DATE_FRAME[1] and this_obj[2] == DATE_FRAME[2]:
            return True
        return False
    elif type(this_obj) == str:
        if this_obj == DATE_FRAME[2]:
            return True
        return False
    return False

def get_fnguide(code):
    get_param = {
        'pGB':1,
        'gicode':'A%s'%(code),
        'cID':'',
        'MenuYn':'Y',
        'ReportGB':'',
        'NewMenuID':101,
        'stkGb':701,
    }
    get_param = parse.urlencode(get_param)
    url="http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?%s"%(get_param)
    try:
        tables = pd.read_html(url, header=0)
    except:
        exceptional_table = []
        return(exceptional_table)
    return(tables)


def get_equity_capital(code):
    try:
        annual = get_fnguide(code)[11] # 연결-연간 재무제표
        if check_data_date(annual.iloc[0][5]):
            return annual.iloc[10][5]
        return 0
    except:
        return 0

def get_roe(code):
    annual = get_fnguide(code)[11] # 연결-연간 재무제표
    try:
        if check_data_date(annual.iloc[0][3:6]):
            return annual.iloc[18][3:6].tolist() # 최근 3개년 확정 ROE
    except:
        return 0

def get_roe_weighted_mean(code):
    roes = get_roe(code)
    if roes == 0:
        return 0
    sum = 0
    for i in range(0,len(roes)):
        sum += float(roes[i]) * (i + 1)
    return sum / 6

def get_ec_and_roe(code):
    try:
        annual = get_fnguide(code)[11] # 연결-연간 재무제표
        dates_list = list(annual.iloc[0][3:6])
        if check_data_date(dates_list):
            EC = annual.iloc[10][5] # 자기자본
            roes = annual.iloc[18][3:6].tolist() # 최근 3개년 확정 ROE
        else:
            print("check data fail")
            roes = 0
            EC = 0

        if roes == 0:
            return EC, roes
        sum = 0
        for i in range(0,len(roes)):
            sum += float(roes[i]) * (i + 1)
        ROE3y = round(sum / 6, 3)
    except:
        EC = 0
        ROE3y = 0
    return EC, ROE3y

def get_intrinsic_value(ec, roe, stocks_num, dc_rate):
    ec = ec + '00000000' # 단위 - 억원
    ec = float(ec)
    roe = float(roe)
    stocks_num = float(stocks_num)
    dc_rate = float(dc_rate)
    corp_value = ec + (ec * (roe - dc_rate) / discount_rate)
    X = int(round(corp_value / stocks_num, -1))
    return X

def get_stocks_num_and_price(code):
    try:
        general = get_fnguide(code)[0]
        stocknum = general.iloc[5][1]
        stocknum = int(stocknum[:stocknum.index('/')].replace(',', ''))
        curr_price = general.columns[1]
        curr_price = int(curr_price[:curr_price.index('/')].replace(',',''))
        return stocknum, curr_price
    except:
        return 0, 0

lock = threading.Lock()
semaphore = threading.BoundedSemaphore(10) # 최대 쓰레드 수 10개

def iterate_stocksDB(i, EC, ROE3y, stock_num, curr_price, intrinsic_value):
    lock.acquire() # 데이터 오염 방지 / 한 파일에 동시에 접근할 때 안전성 확보
    try:
        sheet.cell(i, 7).value = EC
        sheet.cell(i, 8).value = ROE3y
        sheet.cell(i, 9).value = stock_num
        sheet.cell(i, 10).value = curr_price
        sheet.cell(i,11).value = intrinsic_value
        sheet.cell(i,12).value = round(intrinsic_value / curr_price, 2)
        print(i, sheet.cell(i,1).value, '\n', 
                "Equity Capital :", EC, '\n',
                "ROE :", ROE3y, '\n',
                "Num of Stocks :", stock_num, '\n',
                "Current price :", curr_price, '\n',
                "Intrinsic Value :", intrinsic_value, '\n',
                "Potential X:", round(intrinsic_value / curr_price, 2))
    except:
        print("Error")
        sheet.cell(i,11).value = 0
        sheet.cell(i,12).value = 0
    workbook.save(STOCKS_DB_LOC)
    lock.release()

def pass_data(i, EC, ROE3y, stock_num, curr_price, intrinsic_value):
    iterate_stocksDB(i, EC, ROE3y, stock_num, curr_price, intrinsic_value)
    return

def get_data(i):
    semaphore.acquire() # Thread 수 제한
    try:
        if len(str(sheet.cell(i,2).value)) >= 6:
            code = str(sheet.cell(i,2).value)
            EC, ROE3y = get_ec_and_roe(code)
            stock_num, curr_price = get_stocks_num_and_price(code)
            discount_rate = getDiscountRate(KIS_LINK, DISCOUNT_RATE_LOC)
            intrinsic_value = get_intrinsic_value(EC, ROE3y, stock_num, discount_rate)
            pass_data(i, EC, ROE3y, stock_num, curr_price, intrinsic_value)
    except:
        print("Get data Error")
        pass_data(i, 0, 0, 0, 0, 0)
    semaphore.release()
    return

if __name__ == '__main__':
    threads = []
    for i in range(2, 4):
        t = threading.Thread(target=get_data, args=(i, ))
        t.start()
        threads.append(t)

    for thread in threads:
        thread.join()

workbook.close()

#2023/1/8
